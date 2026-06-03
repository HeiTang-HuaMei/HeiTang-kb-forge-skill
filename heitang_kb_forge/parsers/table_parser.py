import csv
from pathlib import Path
from typing import Iterable


def parse_csv(path: Path) -> str:
    return _parse_delimited(path, delimiter=",")


def parse_tsv(path: Path) -> str:
    return _parse_delimited(path, delimiter="\t")


def parse_xlsx(path: Path) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("XLSX dependency is not installed. Install with: pip install -e .") from exc

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise RuntimeError(f"XLSX parsing failed for {path}: {exc}") from exc

    paragraphs: list[str] = []
    for worksheet in workbook.worksheets:
        rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
        paragraphs.extend(_format_table_rows(rows, sheet_name=worksheet.title))
    workbook.close()
    return "\n\n".join(paragraphs)


def _parse_delimited(path: Path, delimiter: str) -> str:
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as file:
            rows = [row for row in csv.reader(file, delimiter=delimiter)]
    except Exception as exc:
        raise RuntimeError(f"Delimited table parsing failed for {path}: {exc}") from exc

    return "\n\n".join(_format_table_rows(rows))


def _format_table_rows(rows: list[list[object]], sheet_name: str | None = None) -> list[str]:
    if not rows:
        return []

    width = max((len(row) for row in rows), default=0)
    if width == 0:
        return []

    first_row = _pad_row(rows[0], width)
    if _row_has_values(first_row):
        headers = _normalize_headers(first_row)
        data_rows = rows[1:]
        row_offset = 2
    else:
        headers = _normalize_headers([])
        data_rows = rows
        row_offset = 1

    headers = _pad_headers(headers, width)
    paragraphs: list[str] = []
    for index, row in enumerate(data_rows, start=row_offset):
        padded = _pad_row(row, width)
        if not _row_has_values(padded):
            continue
        fields = [
            f"{headers[column_index]}: {_stringify(value)}"
            for column_index, value in enumerate(padded)
            if _stringify(value)
        ]
        if not fields:
            continue
        prefix = f"Row {index}."
        if sheet_name is not None:
            prefix = f"Sheet: {sheet_name}. {prefix}"
        paragraphs.append(f"{prefix} {'. '.join(fields)}.")

    return paragraphs


def _normalize_headers(row: Iterable[object]) -> list[str]:
    headers: list[str] = []
    seen: dict[str, int] = {}
    for index, value in enumerate(row):
        header = _stringify(value) or _column_name(index)
        count = seen.get(header, 0) + 1
        seen[header] = count
        headers.append(header if count == 1 else f"{header} {count}")
    return headers


def _pad_headers(headers: list[str], width: int) -> list[str]:
    return headers + [_column_name(index) for index in range(len(headers), width)]


def _pad_row(row: list[object], width: int) -> list[object]:
    return row + [""] * (width - len(row))


def _row_has_values(row: Iterable[object]) -> bool:
    return any(_stringify(value) for value in row)


def _stringify(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _column_name(index: int) -> str:
    name = ""
    number = index + 1
    while number:
        number, remainder = divmod(number - 1, 26)
        name = chr(65 + remainder) + name
    return f"Column {name}"
