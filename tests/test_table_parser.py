import json

from openpyxl import Workbook
from typer.testing import CliRunner

from heitang_kb_forge.cli import app
from heitang_kb_forge.parsers.table_parser import parse_csv, parse_tsv, parse_xlsx


STANDARD_PACKAGE_FILES = {
    "chunks.jsonl",
    "cards.jsonl",
    "qa_pairs.jsonl",
    "glossary.jsonl",
    "manifest.json",
    "ingest_report.md",
    "quality_report.json",
}


def test_csv_parser_formats_rows_with_header(tmp_path):
    path = tmp_path / "products.csv"
    path.write_text("书名,作者,ISBN\n产品经理入门,张三,123456\n", encoding="utf-8")

    text = parse_csv(path)

    assert "Row 2. 书名: 产品经理入门. 作者: 张三. ISBN: 123456." in text


def test_csv_parser_handles_utf8_sig_and_skips_empty_rows(tmp_path):
    path = tmp_path / "products.csv"
    path.write_text("\ufeffName,Price\n\nCourse,99\n,,\n", encoding="utf-8")

    text = parse_csv(path)

    assert "Row 3. Name: Course. Price: 99." in text
    assert "Row 2" not in text
    assert "Row 4" not in text


def test_tsv_parser_formats_tab_delimited_rows(tmp_path):
    path = tmp_path / "products.tsv"
    path.write_text("Name\tPrice\nCourse\t99\n", encoding="utf-8")

    text = parse_tsv(path)

    assert "Row 2. Name: Course. Price: 99." in text


def test_xlsx_parser_formats_single_sheet_rows(tmp_path):
    path = tmp_path / "products.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "商品列表"
    worksheet.append(["书名", "作者", "定价"])
    worksheet.append(["产品经理入门", "张三", 59])
    workbook.save(path)

    text = parse_xlsx(path)

    assert "Sheet: 商品列表. Row 2. 书名: 产品经理入门. 作者: 张三. 定价: 59." in text


def test_xlsx_parser_preserves_multiple_sheet_names_and_skips_empty_rows(tmp_path):
    path = tmp_path / "products.xlsx"
    workbook = Workbook()
    first = workbook.active
    first.title = "商品"
    first.append(["Name"])
    first.append([])
    first.append(["Course"])
    second = workbook.create_sheet("FAQ")
    second.append(["Question", "Answer"])
    second.append(["What is it?", "A table fixture"])
    workbook.save(path)

    text = parse_xlsx(path)

    assert "Sheet: 商品. Row 3. Name: Course." in text
    assert "Sheet: FAQ. Row 2. Question: What is it?. Answer: A table fixture." in text
    assert "Row 2. Name" not in text


def test_table_parser_uses_column_names_for_empty_headers_and_suffixes_duplicates(tmp_path):
    path = tmp_path / "headers.csv"
    path.write_text("Name,,Name\nCourse,99,Duplicate\n", encoding="utf-8")

    text = parse_csv(path)

    assert "Name: Course" in text
    assert "Column B: 99" in text
    assert "Name 2: Duplicate" in text


def test_table_parser_uses_column_names_when_first_row_is_empty(tmp_path):
    path = tmp_path / "no_header.csv"
    path.write_text(",,\nCourse,99\n", encoding="utf-8")

    text = parse_csv(path)

    assert "Row 2. Column A: Course. Column B: 99." in text


def test_build_processes_csv_and_writes_standard_package(tmp_path):
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    input_dir.mkdir()
    (input_dir / "products.csv").write_text("Name,Price\nKB Forge CSV Fixture,99\n", encoding="utf-8")

    result = CliRunner().invoke(app, ["build", "--input", str(input_dir), "--output", str(output_dir)])

    assert result.exit_code == 0, result.output
    assert {path.name for path in output_dir.iterdir()} == STANDARD_PACKAGE_FILES
    assert "KB Forge CSV Fixture" in _read_chunk_text(output_dir)


def test_build_processes_xlsx_and_writes_standard_package(tmp_path):
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    input_dir.mkdir()
    _write_xlsx(input_dir / "products.xlsx", "KB Forge XLSX Fixture")

    result = CliRunner().invoke(app, ["build", "--input", str(input_dir), "--output", str(output_dir)])

    assert result.exit_code == 0, result.output
    assert {path.name for path in output_dir.iterdir()} == STANDARD_PACKAGE_FILES
    assert "KB Forge XLSX Fixture" in _read_chunk_text(output_dir)


def test_batch_processes_csv_and_xlsx(tmp_path):
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    input_dir.mkdir()
    (input_dir / "001_table.csv").write_text("Name\nKB Forge CSV Fixture\n", encoding="utf-8")
    _write_xlsx(input_dir / "002_workbook.xlsx", "KB Forge XLSX Fixture")

    result = CliRunner().invoke(app, ["batch", "--input", str(input_dir), "--output", str(output_dir)])

    assert result.exit_code == 0, result.output
    assert {path.name for path in (output_dir / "001_table").iterdir()} == STANDARD_PACKAGE_FILES
    assert {path.name for path in (output_dir / "002_workbook").iterdir()} == STANDARD_PACKAGE_FILES


def test_batch_table_failure_does_not_block_text_file(monkeypatch, tmp_path):
    import heitang_kb_forge.cli as cli

    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    input_dir.mkdir()
    (input_dir / "001_text.md").write_text("Text fixture", encoding="utf-8")
    (input_dir / "002_table.csv").write_text("Name\nBroken\n", encoding="utf-8")
    monkeypatch.setitem(cli.PARSERS, ".csv", lambda path: (_ for _ in ()).throw(RuntimeError("CSV parsing failed")))

    result = CliRunner().invoke(app, ["batch", "--input", str(input_dir), "--output", str(output_dir)])

    assert result.exit_code == 0, result.output
    manifest = json.loads((output_dir / "batch_manifest.json").read_text(encoding="utf-8"))
    assert manifest["succeeded"] == 1
    assert manifest["failed"] == 1
    items = {item["sequence_id"]: item for item in manifest["items"]}
    assert items["001"]["status"] == "success"
    assert items["002"]["status"] == "failed"
    assert "CSV parsing failed" in items["002"]["error"]


def test_merge_combines_csv_xlsx_and_markdown(tmp_path):
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    input_dir.mkdir()
    (input_dir / "001_notes.md").write_text("KB Forge Markdown Fixture", encoding="utf-8")
    (input_dir / "001_table.csv").write_text("Name\nKB Forge CSV Fixture\n", encoding="utf-8")
    _write_xlsx(input_dir / "001_workbook.xlsx", "KB Forge XLSX Fixture")

    result = CliRunner().invoke(
        app,
        ["batch", "--input", str(input_dir), "--output", str(output_dir), "--merge-same-sequence"],
    )

    assert result.exit_code == 0, result.output
    assert {path.name for path in (output_dir / "001").iterdir()} == STANDARD_PACKAGE_FILES
    chunk_text = _read_chunk_text(output_dir / "001")
    assert "KB Forge Markdown Fixture" in chunk_text
    assert "KB Forge CSV Fixture" in chunk_text
    assert "KB Forge XLSX Fixture" in chunk_text


def _write_xlsx(path, value):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Data"
    worksheet.append(["Name"])
    worksheet.append([value])
    workbook.save(path)


def _read_chunk_text(output_dir):
    return "\n".join(
        json.loads(line)["text"] for line in (output_dir / "chunks.jsonl").read_text(encoding="utf-8").splitlines()
    )
